import os
import pandas as pd
import matplotlib.pyplot as plt
from sqlalchemy import create_engine
from openpyxl.formatting.rule import ColorScaleRule
import plotly.express as px
import pandas as pd

# ---------- DB CONNECTION ----------
engine = create_engine(
    "postgresql+psycopg2://postgres:balnur06@localhost:5432/postgres",
    client_encoding="utf8"
)

# ---------- HELPERS ----------
def run_query(query):
    with engine.connect() as conn:
        raw_conn = conn.connection  # Get the raw DBAPI connection
        return pd.read_sql_query(query, raw_conn)

def save_chart(fig, filename, description, rows):
    os.makedirs("charts", exist_ok=True)
    fig.savefig(f"charts/{filename}", bbox_inches="tight")
    print(f"Saved {filename}: {rows} rows → {description}")

plt.style.use("seaborn-v0_8")

# ---------- 1. PIE CHART: Skaters per Region ----------
query_pie = """
SELECT e."Region", COUNT(s."skater_detail_id") AS skater_count
FROM scraped_skater_details s
JOIN scraped_event_details e 
    ON s."Event_Index" = e."Event_Index"
GROUP BY e."Region";
"""
df_pie = run_query(query_pie)

if not df_pie.empty:
    fig1, ax1 = plt.subplots()
    df_pie.set_index("Region")["skater_count"].plot.pie(
        autopct="%1.1f%%", ax=ax1, startangle=90, cmap="Set3"
    )
    ax1.set_ylabel("")
    ax1.set_title("Распределение участников по регионам")
    save_chart(fig1, "pie_skaters_by_region.png", "Pie chart of skaters by region", len(df_pie))

# ---------- 2. BAR CHART: Avg skaters per Level ----------
query_bar = """
SELECT e."Level", COUNT(s."skater_detail_id") AS skater_count
FROM scraped_skater_details s
JOIN scraped_event_details e 
    ON s."Event_Index" = e."Event_Index"
GROUP BY e."Level";
"""
df_bar = run_query(query_bar)

if not df_bar.empty:
    fig2, ax2 = plt.subplots()
    df_bar.set_index("Level")["skater_count"].plot.bar(ax=ax2, color="skyblue")
    ax2.set_title("Среднее число участников по уровню соревнования")
    ax2.set_xlabel("Level")
    ax2.set_ylabel("Число участников")
    save_chart(fig2, "bar_avg_skaters_per_level.png", "Bar chart of avg skaters per level", len(df_bar))

# ---------- 3. HORIZONTAL BAR: Officials by Country ----------
query_barh = """
SELECT o."Official_StateOrCountry", COUNT(o."Event_Index") AS officials_count
FROM scraped_official_details o
GROUP BY o."Official_StateOrCountry"
ORDER BY officials_count DESC
LIMIT 10;
"""
df_barh = run_query(query_barh)

if not df_barh.empty:
    fig3, ax3 = plt.subplots()
    df_barh.set_index("Official_StateOrCountry")["officials_count"].plot.barh(ax=ax3, color="lightgreen")
    ax3.set_title("Количество судей по странам (Топ 10)")
    ax3.set_xlabel("Количество")
    save_chart(fig3, "barh_officials_by_country.png", "Horizontal bar chart of officials", len(df_barh))

# ---------- 4. LINE CHART: Events Over Time ----------
query_line = """
SELECT "Event_Date"
FROM scraped_event_details
WHERE "Event_Date" IS NOT NULL;
"""

df_line = run_query(query_line)

# Convert Event_Date to datetime
df_line["Event_Date"] = pd.to_datetime(df_line["Event_Date"], errors="coerce")
df_line = df_line.dropna(subset=["Event_Date"])

# Group by day and count events
df_daily = df_line.groupby("Event_Date").size().reset_index(name="event_count")

# ---------- Animation ----------
query_bar_anim = """
SELECT 
    e."Level", 
    EXTRACT(YEAR FROM e."Event_Date") AS year,
    COUNT(s."skater_detail_id") AS skater_count
FROM scraped_skater_details s
JOIN scraped_event_details e 
    ON s."Event_Index" = e."Event_Index"
GROUP BY e."Level", year
ORDER BY year, e."Level";
"""
df_bar_anim = run_query(query_bar_anim)

# Преобразуем year в int
df_bar_anim['year'] = df_bar_anim['year'].astype(int)

if not df_bar_anim.empty:
    fig = px.bar(
        df_bar_anim, 
        x='Level', 
        y='skater_count', 
        color='Level',
        animation_frame='year',
        range_y=[0, df_bar_anim['skater_count'].max() + 5],
        title='Число участников по уровням соревнования с анимацией по годам',
        labels={'skater_count': 'Число участников', 'Level': 'Уровень'}
    )
    fig.write_html("events1_slider.html")
    print("График сохранён в файл events1_slider.html")


# ---------- 5. HISTOGRAM: GOE Score Distribution ----------
query_hist = """
SELECT t."GOE"
FROM scraped_technical_scores t
WHERE t."GOE" IS NOT NULL;
"""
df_hist = run_query(query_hist)

if not df_hist.empty:
    fig5, ax5 = plt.subplots()
    df_hist["GOE"].plot.hist(bins=15, ax=ax5, color="plum")
    ax5.set_title("Распределение GOE баллов")
    ax5.set_xlabel("GOE")
    ax5.set_ylabel("Частота")
    save_chart(fig5, "hist_goe_scores.png", "Histogram of GOE scores", len(df_hist))

# ---------- 6. SCATTER: Technical vs Component by Level ----------
query_scatter = """
SELECT e."Level", 
       AVG(t."PanelScores_Technical") AS avg_technical, 
       AVG(c."PanelScores_PC") AS avg_component
FROM scraped_technical_scores t
JOIN scraped_event_details e 
    ON t."Event_Index" = e."Event_Index"
JOIN scraped_component_scores c 
    ON t."Event_Index" = c."Event_Index"
GROUP BY e."Level";
"""
df_scatter = run_query(query_scatter)

if not df_scatter.empty:
    fig6, ax6 = plt.subplots()
    ax6.scatter(df_scatter["avg_technical"], df_scatter["avg_component"], color="orange")
    for i, level in enumerate(df_scatter["Level"]):
        ax6.text(df_scatter["avg_technical"][i], df_scatter["avg_component"][i], level)
    ax6.set_title("Технические vs Компонентные баллы по уровню")
    ax6.set_xlabel("Avg Technical Score")
    ax6.set_ylabel("Avg Component Score")
    save_chart(fig6, "scatter_tech_vs_component.png", "Scatter plot of scores by level", len(df_scatter))

# ---------- EXPORT TO EXCEL ----------
def export_to_excel(dataframes_dict, filename):
    os.makedirs("exports", exist_ok=True)
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        total_rows = 0
        for sheet, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
            total_rows += len(df)

        wb = writer.book
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.freeze_panes = "B2"
            ws.auto_filter.ref = ws.dimensions
            for col in ws.iter_cols(min_col=2, max_col=ws.max_column,
                                    min_row=2, max_row=ws.max_row):
                if len(col) == 0:
                    continue
                rng = f"{col[0].column_letter}2:{col[0].column_letter}{ws.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(rng, rule)

    print(f"Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")

export_to_excel(
    {
        "Pie_Region": df_pie,
        "Bar_Level": df_bar,
        "BarH_Officials": df_barh,
        "Line_Events": df_line,
        "Histogram_GOE": df_hist,
        "Scatter_Scores": df_scatter
    },
    "exports/skaters_report.xlsx"
)
